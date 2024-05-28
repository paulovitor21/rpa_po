import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Cria um novo workbook e seleciona a sheet ativa
wb = openpyxl.Workbook()
ws = wb.active

# Define os dados da tabela
data = [
    ["NCM/ HS CODE", "PART Nº", "DESCRIPTION", "QTY", "US$ UNIT", "US$ TOTAL"],
    ["85444200#000", "GGLZ241029405", "VACUUM HOLDER KSLD100S H420 B850P", 1, "$ 2.072,00", "$ 2.072,00"],
    ["84719012#000", "GGLZ241029406", "Amount", 1, "", "$ 2.072,00"]
]

# Preenche os dados na planilha
for row in data:
    ws.append(row)

# Ajusta a largura das colunas para melhor visualização
column_widths = [15, 18, 60, 5, 10, 10]
for i, column_width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = column_width

# Estilo para o cabeçalho
header_fill = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
header_alignment = Alignment(horizontal="center", vertical="center")

# Aplica o estilo ao cabeçalho
for col in range(1, len(data[0]) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Estilo para o corpo da tabela
body_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
body_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Aplica o estilo ao corpo da tabela
for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=6):
    for cell in row:
        cell.alignment = body_alignment
        cell.border = thin_border
        if cell.column in [4, 5, 6]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if cell.column == 5 and cell.row == 2:
            cell.fill = body_fill

# Salva o arquivo Excel
wb.save("tabela.xlsx")
